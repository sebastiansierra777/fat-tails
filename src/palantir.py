# %%----------------------------------------------------------------------
# PALANTIR LOG-RETURNS
import pandas as pd
import numpy as np
from pathlib import Path

from patsy import desc
from scipy.stats import skew, kurtosis
import openpyxl

# Establish Excel Path
excel_path = Path(r"C:\Users\sebas\Documents\Python\Python Projects\fat_tails\data\processed\fat_tails.xlsx")

# Load the data
df = pd.read_excel(excel_path, sheet_name="palantir")

# Parse and sort dates
df["Date"] = pd.to_datetime(df["Date"]).dt.date
df = df.sort_values(by=["Date"])

# Use closing price, simply known as "Price" in the file columns
price_col = "Price"

# Compute log returns
df["log_return"] = np.log(df[price_col] / df[price_col].shift(1))

# Convert log returns to percentages
df["log_return_percentage"] = df["log_return"] * 100

# Drop the first row, which will be a NaN, since Jan 2 can't be compared
df = df.dropna(subset=["log_return"])

# ------------STATS-----------------------------------------------------

# Compute descriptive stats
describe_stats = df["log_return"].describe()

# Skewness & Excess Kurtosis
skewness = skew(df["log_return"], bias=False)
excess_kurt = kurtosis(df["log_return"], bias=False)

# Annualized return (from daily log returns)
trading_days = 252
mean_daily_log_ret = df["log_return"].mean()
annualized_return = np.exp(mean_daily_log_ret * trading_days) - 1

# ------------PRINT PREVIEWS FOR THE USER--------------------------------

print("\n================ PALANTIR LOG-RETURNS PREVIEW ================\n")
print(df[["Date", price_col, "log_return", "log_return_percentage"]].head())

print("\n================ DESCRIPTIVE STATS ===========================\n")
print(describe_stats)

print("\n================ SKEWNESS & KURTOSIS =========================\n")
print(f"Skewness:          {skewness:.6f}")
print(f"Excess Kurtosis:   {excess_kurt:.6f}")

print("\n================ ANNUALIZED RETURN ===========================\n")
print(f"Annualized Return: {annualized_return:.4%}\n")

#----------------------EXPORTS----------------------------------------------

# Write the DataFrame to an existing Excel file
with pd.ExcelWriter(
    excel_path,
    engine="openpyxl",
    mode="a",
    if_sheet_exists="replace",
) as writer:
    df.to_excel(writer, sheet_name="palantir_log", index=False)

# -----------------APPEND STATS TABLE BELOW THE DATA----------------------

# Load workbook / sheet
wb = openpyxl.load_workbook(excel_path)
ws = wb["palantir_log"]

# Find the first empty row below the dataframe
start_row = ws.max_row + 2 # two rows of space
start_col =  2 # 1 column of space

# List of rows (label, value)
rows = [
    ("count", describe_stats["count"]),
    ("mean", describe_stats["mean"]),
    ("std", describe_stats["std"]),
    ("min", describe_stats["min"]),
    ("25%", describe_stats["25%"]),
    ("50%", describe_stats["50%"]),
    ("75%", describe_stats["75%"]),
    ("max", describe_stats["max"]),
    ("skewness", skewness),
    ("excess_kurtosis", excess_kurt),
    ("annualized_return", annualized_return),
]

# Write the stats table
for i, (label, value) in enumerate(rows, start=start_row + 1):
    ws.cell(row=i, column=start_col, value=label)
    ws.cell(row=i, column=start_col + 1, value=value)

# Save the workbook
wb.save(excel_path)

print("\nStats table successfully added below the data")









